"""
Excel reader - loads company names and career URLs from Excel files.
"""

import logging
from pathlib import Path
from typing import List, Dict

logger = logging.getLogger(__name__)


def load_companies_from_excel(file_path: str) -> List[Dict]:
    """
    Read companies from an Excel file.
    Expects columns: Company Name, Career Portal Link (or similar).
    Auto-detects column names.
    Category is extracted from the filename (e.g., 'Semiconductor_Companies_Careers.xlsx' → 'Semiconductor').
    """
    try:
        import openpyxl
    except ImportError:
        import subprocess, sys
        subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
        import openpyxl

    path = Path(file_path)
    if not path.exists():
        logger.error(f"File not found: {file_path}")
        return []

    # Extract category from filename (first word before '_')
    stem = path.stem  # e.g., "Semiconductor_Companies_Careers_Updated"
    category = stem.split("_")[0].strip()
    if not category:
        category = "Other"

    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    ws = wb.active

    # Read headers
    headers = []
    for cell in ws[1]:
        val = str(cell.value or "").strip().lower()
        headers.append(val)

    # Find company name and URL columns
    name_col = None
    url_col = None
    for i, h in enumerate(headers):
        if any(k in h for k in ["company", "name", "organization"]):
            name_col = i
        if any(k in h for k in ["career", "link", "url", "portal", "page"]):
            url_col = i

    # Fallback: assume first column is name, second is URL
    if name_col is None:
        name_col = 0
    if url_col is None and len(headers) > 1:
        url_col = 1

    companies = []
    seen = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[name_col]:
            continue
        name = str(row[name_col]).strip()
        url = str(row[url_col]).strip() if url_col is not None and row[url_col] else ""

        # Skip duplicates
        key = name.lower()
        if key in seen:
            continue
        seen.add(key)

        # Skip invalid URLs
        if url and not url.startswith("http"):
            url = ""

        if name and url:
            companies.append({"name": name, "career_url": url, "category": category})
        elif name:
            logger.warning(f"No career URL for: {name}")

    wb.close()
    logger.info(f"Loaded {len(companies)} companies from {file_path} (category: {category})")
    return companies


def load_companies_from_multiple_files(file_paths: List[str]) -> List[Dict]:
    """Load and merge companies from multiple Excel files."""
    all_companies = []
    seen = set()
    for fp in file_paths:
        for company in load_companies_from_excel(fp):
            key = company["name"].lower()
            if key not in seen:
                seen.add(key)
                all_companies.append(company)
    logger.info(f"Total unique companies from {len(file_paths)} files: {len(all_companies)}")
    return all_companies
